from __future__ import annotations

import json
import sys
from pathlib import Path


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python3 scripts/import_gesp_trial_users.py /path/to/trial-users.xlsx")

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("openpyxl is required to read .xlsx files") from error

    source = Path(sys.argv[1]).expanduser().resolve()
    output = Path(__file__).resolve().parents[1] / "data" / "gesp_trial_users.json"
    workbook = load_workbook(source, data_only=True)
    sheet = workbook.active
    headers = [str(value).strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    users = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {headers[index]: row[index] for index in range(len(headers))}
        username = str(data.get("用户名") or "").strip()
        if not username:
            continue
        users.append({
            "username": username,
            "password": str(data.get("密码") or "").strip(),
            "name": str(data.get("姓名") or "").strip(),
            "language": str(data.get("语言") or "").strip(),
            "level": str(data.get("等级") or "").strip(),
        })

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps({"source": source.name, "count": len(users), "items": users}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Imported {len(users)} GESP trial users -> {output}")


if __name__ == "__main__":
    main()
